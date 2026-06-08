"""SkillOpt-Sleep — SpreadsheetBench execution + scoring (faithful port).

The spreadsheet task is NOT text Q&A: the agent must write openpyxl Python code
that reads an input workbook, applies the requested manipulation, and saves an
output workbook; the official evaluator then compares the answer cells to a
golden workbook. This module vendors the intern's executor + cell-value compare
so the plugin scores spreadsheet EXACTLY as the research repo does.

Source: skillopt/envs/spreadsheetbench/{executor.py,evaluator.py}
"""
from __future__ import annotations

import datetime
import glob
import os
import re
import subprocess
import sys
import tempfile
import textwrap
from typing import Tuple


# ── code execution (port of executor.run_generated_code) ──────────────────────

_RUNNER = textwrap.dedent(
    """
    import os, sys, traceback
    INPUT_PATH = {input_path!r}
    OUTPUT_PATH = {output_path!r}
    try:
    {user_code_indented}
    except Exception:
        traceback.print_exc()
        sys.exit(2)
    """
)
_PATH_ASSIGN = re.compile(r'^\s*(INPUT_PATH|OUTPUT_PATH)\s*=\s*.+$', re.MULTILINE)


def _extract_code(text: str) -> str:
    """Pull a python code block out of the model response (```python ... ```)."""
    m = re.findall(r"```(?:python)?\s*(.*?)```", text or "", re.DOTALL | re.IGNORECASE)
    if m:
        return max(m, key=len).strip()
    return (text or "").strip()


def run_generated_code(code: str, input_path: str, output_path: str, timeout: int = 120) -> Tuple[bool, str]:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    cleaned = _PATH_ASSIGN.sub("", code)
    script = _RUNNER.format(input_path=input_path, output_path=output_path,
                            user_code_indented=textwrap.indent(cleaned, "    "))
    with tempfile.NamedTemporaryFile("w", suffix=".py", delete=False) as f:
        f.write(script)
        tmp = f.name
    try:
        proc = subprocess.run([sys.executable, tmp], capture_output=True, text=True, timeout=timeout)
        if proc.returncode != 0:
            return False, (proc.stderr or "exec failed")[:200]
        return os.path.exists(output_path), ""
    except Exception as e:  # noqa: BLE001
        return False, f"{type(e).__name__}: {e}"
    finally:
        try:
            os.unlink(tmp)
        except Exception:
            pass


# ── cell-value compare (port of evaluator) ────────────────────────────────────

def _dt_to_float(dt: datetime.datetime) -> float:
    return (dt - datetime.datetime(1899, 12, 30)).days


def _transform(v):
    import datetime as _d
    if isinstance(v, bool):
        return round(float(v), 2)
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, str):
        try:
            return round(float(v), 2)
        except ValueError:
            return v
    if isinstance(v, _d.datetime):
        return int(_dt_to_float(v))
    if isinstance(v, _d.time):
        return str(v).split(".")[0]
    return v


def _eq(a, b) -> bool:
    if (a is None or a == "") and (b is None or b == ""):
        return True
    ta, tb = _transform(a), _transform(b)
    if type(ta) is not type(tb):
        # numeric cross-type tolerance
        try:
            return round(float(ta), 2) == round(float(tb), 2)
        except (ValueError, TypeError):
            return False
    return ta == tb


def _iter_coords(ws, cell_range: str):
    """Yield cell coordinates for a range, robust to single-cell / single-row."""
    sel = ws[cell_range]
    # openpyxl returns: a Cell (single), a tuple of Cells (single row/col), or a
    # tuple of tuples of Cells (rectangular range).
    if hasattr(sel, "coordinate"):           # single Cell
        yield sel.coordinate
        return
    for row in sel:
        if hasattr(row, "coordinate"):       # single row/col -> row is a Cell
            yield row.coordinate
        else:
            for cell in row:
                yield cell.coordinate


def _cell_range(ws_gt, ws_pr, cell_range: str) -> bool:
    for coord in _iter_coords(ws_gt, cell_range):
        if not _eq(ws_gt[coord].value, ws_pr[coord].value):
            return False
    return True


def compare_workbooks(gold: str, pred: str, answer_position: str, answer_sheet: str = "") -> bool:
    import openpyxl
    if not os.path.exists(pred):
        return False
    try:
        wg = openpyxl.load_workbook(gold, data_only=True)
        wp = openpyxl.load_workbook(pred, data_only=True)
    except Exception:
        return False
    try:
        for scr in (answer_position or "").split(","):
            scr = scr.strip()
            if not scr:
                continue
            if "!" in scr:
                sheet, rng = scr.split("!", 1)
                sheet = sheet.strip().strip("'\"")
            else:
                sheet = answer_sheet or wg.sheetnames[0]
                rng = scr
            rng = rng.strip().strip("'\"")
            try:
                wsg = wg[sheet] if sheet in wg.sheetnames else wg[wg.sheetnames[0]]
                wsp = wp[sheet] if sheet in wp.sheetnames else wp[wp.sheetnames[0]]
            except Exception:
                return False
            if not _cell_range(wsg, wsp, rng):
                return False
        return True
    finally:
        wg.close(); wp.close()


# ── locate the task workbooks ─────────────────────────────────────────────────

def _find_files(task_dir: str) -> Tuple[str, str]:
    """Return (input_xlsx, golden_xlsx) for a verified_400 task dir."""
    inits = sorted(glob.glob(os.path.join(task_dir, "*_init.xlsx")))
    if inits:
        ip = inits[0]
        return ip, ip.replace("_init.xlsx", "_golden.xlsx")
    inputs = sorted(glob.glob(os.path.join(task_dir, "*_input.xlsx")))
    if inputs:
        ip = inputs[0]
        return ip, ip.replace("_input.xlsx", "_answer.xlsx")
    bi = os.path.join(task_dir, "initial.xlsx")
    if os.path.exists(bi):
        return bi, os.path.join(task_dir, "golden.xlsx")
    return "", ""


def score_case(response: str, item: dict, data_root: str) -> Tuple[float, float, str]:
    sp = item.get("spreadsheet_path", "")
    task_dir = os.path.join(data_root, sp)
    inp, gold = _find_files(task_dir)
    if not inp or not os.path.exists(inp) or not os.path.exists(gold):
        return 0.0, 0.0, f"workbook-missing:{sp}"
    code = _extract_code(response)
    if not code:
        return 0.0, 0.0, "no-code"
    out = tempfile.NamedTemporaryFile(suffix="_out.xlsx", delete=False).name
    try:
        ok, msg = run_generated_code(code, inp, out)
        if not ok:
            return 0.0, 0.0, f"exec:{msg[:60]}"
        passed = compare_workbooks(gold, out, item.get("answer_position", ""),
                                   item.get("answer_sheet", ""))
        h = 1.0 if passed else 0.0
        return h, h, ("pass" if passed else "cell-mismatch")
    finally:
        try:
            os.unlink(out)
        except Exception:
            pass
